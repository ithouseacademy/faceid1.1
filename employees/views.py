# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from datetime import date, datetime, time, timedelta
import json
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Count
from .models import ManualEntryPhoto
# AUTHENTICATION IMPORTS
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm, SetPasswordForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User

from .models import Employee, Attendance, MonthlySalary, UserProfile, ManualEntryPhoto, Location, FaceCapture, LateAbsenceRecord, CleanupConfig

from .forms import EmployeeForm, UserCreateForm, UserEditForm

import csv
import io
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.core.files.base import ContentFile
import base64
import uuid





# views.py ga qo'shing
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.files.base import ContentFile
import base64
import json
import uuid
from .models import ManualEntryPhoto

from django.core.paginator import Paginator
from django.db.models import Count
from .models import ManualEntryPhoto


@login_required
def get_employee_photos_api(request, employee_id):
    """API to get photos for a specific employee with attendance type"""
    try:
        employee = get_object_or_404(Employee, id=employee_id)

        # Get photos from ManualEntryPhoto model
        photos = ManualEntryPhoto.objects.filter(employee=employee).order_by('-captured_at')

        photos_data = []
        for photo in photos:
            photos_data.append({
                'id': photo.id,
                'url': photo.photo.url if photo.photo else None,
                'attendance_type': photo.attendance_type,
                'type_display': photo.type_display,
                'captured_at': photo.captured_at.strftime('%Y-%m-%d %H:%M:%S'),
                'ip_address': photo.ip_address,
            })

        return JsonResponse({
            'status': 'success',
            'employee_id': employee_id,
            'employee_name': employee.full_name,
            'photos': photos_data,
            'total_photos': len(photos_data)
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)









def manual_employee_photos(request):
    """Qo'lda kiritilgan xodimlar rasmlari"""
    photos = ManualEntryPhoto.objects.select_related('employee').all()

    # Filterlar
    employee_id = request.GET.get('employee')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if employee_id:
        photos = photos.filter(employee_id=employee_id)

    if date_from:
        photos = photos.filter(captured_at__date__gte=date_from)

    if date_to:
        photos = photos.filter(captured_at__date__lte=date_to)

    # Statistika
    total_photos = photos.count()
    unique_employees = photos.values('employee').distinct().count()
    today_photos = photos.filter(captured_at__date=timezone.now().date()).count()
    unique_ips = photos.exclude(ip_address__isnull=True).values('ip_address').distinct().count()

    # Pagination
    paginator = Paginator(photos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'photos': page_obj,
        'total_photos': total_photos,
        'unique_employees': unique_employees,
        'today_photos': today_photos,
        'unique_ips': unique_ips,
        'employees': Employee.objects.all(),
    }

    return render(request, 'qolda_kiritilgan_xodimlar.html', context)





@csrf_exempt
@require_POST
def save_manual_entry_photo(request):
    """Qo'lda kiritilgan xodimning rasmini saqlash (kelish/chiqish turi bilan)"""
    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
        employee_name = data.get('employee_name', '')
        image_data = data.get('image_data')  # base64 formatda
        attendance_type = data.get('attendance_type', 'in')  # 'in' yoki 'out'
        timestamp = data.get('timestamp', '')

        if not employee_id or not image_data:
            return JsonResponse({'status': 'error', 'message': 'Ma\'lumotlar yetarli emas'})

        # Base64 dan rasmni olish
        format, imgstr = image_data.split(';base64,')
        ext = format.split('/')[-1]

        # Fayl nomi yaratish
        filename = f"manual_{employee_id}_{attendance_type}_{uuid.uuid4().hex}.{ext}"

        # Rasmni saqlash
        photo = ManualEntryPhoto(
            employee_id=employee_id,
            attendance_type=attendance_type,  # TURNI SAQLASH
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        # Rasmni ContentFile orqali saqlash
        photo.photo.save(
            filename,
            ContentFile(base64.b64decode(imgstr)),
            save=True
        )

        return JsonResponse({
            'status': 'success',
            'message': 'Rasm muvaffaqiyatli saqlandi',
            'photo_url': photo.photo.url,
            'attendance_type': attendance_type,
            'type_display': "Kelish" if attendance_type == 'in' else "Chiqish"
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})





















# ============ PERMISSION DECORATORS ============

def admin_only_required(function=None):
    """Faqat adminlar uchun (superuser, is_staff yoki user_type='admin')"""
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and (
            u.is_superuser or
            u.is_staff or
            (hasattr(u, 'profile') and u.profile.user_type == 'admin')
        ),
        login_url='login'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator


def admin_or_hr_required(function=None):
    """Faqat admin yoki HR uchun"""
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and (
            u.is_superuser or
            u.is_staff or
            (hasattr(u, 'profile') and u.profile.user_type in ['admin', 'hr'])
        ),
        login_url='login'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator


def employee_only_required(function=None):
    """Faqat oddiy xodimlar uchun (employee user_type)"""
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and (
            not u.is_superuser and
            not u.is_staff and
            (hasattr(u, 'profile') and u.profile.user_type == 'employee')
        ),
        login_url='login'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator


# ============ AUTHENTICATION VIEWS ============

def login_view(request):
    """Login sahifasi"""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Xush kelibsiz, {user.username}!')
            return redirect('home')
        else:
            messages.error(request, 'Login yoki parol notoʻgʻri!')
    else:
        form = AuthenticationForm()

    return render(request, 'auth/login.html', {'form': form})


def logout_view(request):
    """Logout funksiyasi"""
    logout(request)
    messages.info(request, 'Siz tizimdan chiqdingiz!')
    return redirect('login')


@login_required
def profile_view(request):
    """User profilini ko'rish"""
    user = request.user

    # Agar profile bo'lmasa yaratish
    if not hasattr(user, 'profile'):
        UserProfile.objects.create(user=user, user_type='employee')

    if request.method == 'POST':
        # Asosiy ma'lumotlar
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()

        # Profil ma'lumotlari
        profile = user.profile
        profile.phone = request.POST.get('phone', profile.phone)
        profile.department = request.POST.get('department', profile.department)
        profile.save()

        messages.success(request, 'Profil ma\'lumotlari yangilandi!')
        return redirect('profile')

    return render(request, 'auth/profile.html', {
        'user': user,
        'profile': user.profile,
    })


# ============ ADMIN DASHBOARD (faqat adminlar) ============

@login_required
@admin_only_required
def admin_dashboard(request):
    """Admin dashboard - faqat adminlar uchun"""
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_employees = Employee.objects.filter(is_active=True).count()

    # Adminlar soni
    admin_count = User.objects.filter(
        Q(is_superuser=True) |
        Q(is_staff=True) |
        Q(profile__user_type='admin')
    ).distinct().count()

    # So'nggi foydalanuvchilar
    recent_users = User.objects.all().order_by('-date_joined')[:10]

    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_employees': total_employees,
        'admin_count': admin_count,
        'recent_users': recent_users,
    }
    return render(request, 'admin/dashboard.html', context)


# ============ USER MANAGEMENT (faqat adminlar) ============

@login_required
@admin_only_required
def user_list(request):
    """Userlar ro'yxati - faqat adminlar uchun"""
    users = User.objects.all().order_by('-date_joined')

    # Search funksiyasi
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Filter by user type
    user_type = request.GET.get('type', '')
    if user_type:
        users = users.filter(profile__user_type=user_type)

    # Faqat aktiv foydalanuvchilar
    active_only = request.GET.get('active', '')
    if active_only == 'true':
        users = users.filter(is_active=True)

    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Statistikalar
    admin_count = users.filter(profile__user_type='admin').count()
    hr_count = users.filter(profile__user_type='hr').count()
    manager_count = users.filter(profile__user_type='manager').count()
    employee_count = users.filter(profile__user_type='employee').count()
    superuser_count = users.filter(is_superuser=True).count()
    active_count = users.filter(is_active=True).count()

    context = {
        'users': page_obj,
        'search_query': search_query,
        'user_type': user_type,
        'total_users': users.count(),
        'admin_count': admin_count,
        'hr_count': hr_count,
        'manager_count': manager_count,
        'employee_count': employee_count,
        'superuser_count': superuser_count,
        'active_count': active_count,
        'today': date.today(),
    }
    return render(request, 'admin/user_list.html', context)

@login_required
@admin_only_required
def add_user(request):
    """Yangi user qo'shish - faqat adminlar uchun"""
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'{user.username} foydalanuvchi muvaffaqiyatli yaratildi!')
            return redirect('user_list')
        else:
            messages.error(request, 'Iltimos, xatolarni tuzating!')
    else:
        form = UserCreateForm()

    employees = Employee.objects.filter(is_active=True)

    return render(request, 'admin/user_form.html', {
        'form': form,
        'employees': employees,
        'title': 'Yangi foydalanuvchi qoʻshish',
        'action': 'add',
    })


@login_required
@admin_only_required
def edit_user(request, user_id):
    """Userni tahrirlash - faqat adminlar uchun"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'{user.username} ma\'lumotlari yangilandi!')
            return redirect('user_list')
        else:
            messages.error(request, 'Iltimos, xatolarni tuzating!')
    else:
        form = UserEditForm(instance=user)

    employees = Employee.objects.filter(is_active=True)

    return render(request, 'admin/user_form.html', {
        'form': form,
        'user': user,
        'employees': employees,
        'title': 'Foydalanuvchini tahrirlash',
        'action': 'edit',
    })


@csrf_exempt
@login_required
@admin_only_required
def delete_user(request, user_id):
    """Userni o'chirish - faqat adminlar uchun"""
    if request.method == 'POST':
        try:
            user = get_object_or_404(User, id=user_id)

            # O'zini o'chirishga yo'l qo'ymaslik
            if user == request.user:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Ozingizni ochira olmaysiz'
                })

            username = user.username
            user.delete()

            messages.success(request, f'{username} foydalanuvchi ochirildi!')
            return JsonResponse({
                'status': 'success',
                'message': f'{username} foydalanuvchi ochirildi!'
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Faqat POST sorovi'})


@login_required
@admin_only_required
def change_user_password(request, user_id):
    """User parolini o'zgartirish - faqat adminlar uchun"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'{user.username} paroli muvaffaqiyatli ozgartirildi!')
            return redirect('user_list')
        else:
            messages.error(request, 'Parol notoʻgʻri yoki bir xil emas!')
    else:
        form = SetPasswordForm(user)

    return render(request, 'admin/change_password.html', {
        'form': form,
        'user': user,
        'title': f'{user.username} parolini ozgartirish',
    })


# ============ ASOSIY SAHIFALAR (barcha foydalanuvchilar) ============

def home(request):
    """Asosiy sahifa - login sahifasiga yo'naltiradi"""
    if request.user.is_authenticated:
        return render(request, 'home.html')
    return redirect('login')


@login_required
def dashboard(request):
    """Dashboard - barcha foydalanuvchilar uchun"""
    today = date.today()

    today_checkins = Attendance.objects.filter(date=today, type='in')
    today_checkouts = Attendance.objects.filter(date=today, type='out')

    present_employees = []
    for checkin in today_checkins:
        if not today_checkouts.filter(employee=checkin.employee).exists():
            present_employees.append(checkin.employee)

    all_employees = Employee.objects.filter(is_active=True)
    late_today = today_checkins.filter(status='late')
    day_off_today = today_checkins.filter(status='day_off')

    current_month = today.month
    current_year = today.year
    monthly_salaries = MonthlySalary.objects.filter(year=current_year, month=current_month)

    total_salary = monthly_salaries.aggregate(total=Sum('net_salary'))['total'] or 0
    total_penalty = monthly_salaries.aggregate(total=Sum('total_penalty'))['total'] or 0

    context = {
        'today_checkins': today_checkins,
        'today_checkouts': today_checkouts,
        'present_employees': present_employees,
        'total_employees': all_employees.count(),
        'present_today': today_checkins.count(),
        'checked_out_today': today_checkouts.count(),
        'still_working': len(present_employees),
        'late_today': late_today.count(),
        'day_off_today': day_off_today.count(),
        'all_employees': all_employees,
        'today': today,
        'total_salary': total_salary,
        'total_penalty': total_penalty,
        'active_count': all_employees.count(),
        'absent_today': all_employees.count() - today_checkins.count(),
    }
    return render(request, 'dashboard.html', context)


# ============ XODIMLAR BOSHQARUVI (admin yoki HR) ============

@login_required
@admin_or_hr_required
def employee_list(request):
    """Xodimlar ro'yxati - admin yoki HR uchun"""
    employees = Employee.objects.all().order_by('first_name', 'last_name')
    today = date.today()

    today_checkins = Attendance.objects.filter(date=today, type='in')
    checked_in_ids = list(today_checkins.values_list('employee_id', flat=True))

    # Bugungi jarimalar
    today_penalties = {}
    for checkin in today_checkins.filter(status='late'):
        today_penalties[checkin.employee_id] = checkin.penalty_amount

    # Calculate penalties for each employee
    employees_with_penalties = []
    for employee in employees:
        # Today's penalty
        today_penalty = today_penalties.get(employee.id, 0)

        # Current month penalty details
        penalty_details = employee.get_current_month_penalty_details()

        # Monthly penalty (hamma jarimalarni jamlab)
        monthly_penalty = penalty_details['total_penalty']

        # Get penalty history for last 3 months
        penalty_history = employee.get_monthly_penalty_history(limit=3)

        # Daily penalties for current month
        daily_penalties = penalty_details.get('daily_penalties', [])

        employees_with_penalties.append({
            'employee': employee,
            'today_penalty': today_penalty,
            'monthly_penalty': monthly_penalty,
            'penalty_details': penalty_details,
            'penalty_history': penalty_history,
            'daily_penalties': daily_penalties[:5],  # Faqat oxirgi 5 kunlik jarimalar
            'checked_in_today': employee.id in checked_in_ids
        })

    # Calculate totals
    total_employees = employees.count()
    present_today = today_checkins.count()

    # Oylik jarimalarni jamlab hisoblaymiz
    total_monthly_penalty = 0
    total_today_penalty = 0

    for emp in employees_with_penalties:
        total_monthly_penalty += emp['monthly_penalty']
        total_today_penalty += emp['today_penalty']

    # Kechikgan xodimlar soni
    late_today_count = len([emp for emp in employees_with_penalties if emp['today_penalty'] > 0])

    # Oylik jarimalar soni
    monthly_penalty_count = len([emp for emp in employees_with_penalties if emp['monthly_penalty'] > 0])

    context = {
        'employees_with_penalties': employees_with_penalties,
        'today_checkins': today_checkins,
        'checked_in_ids': checked_in_ids,
        'today_day': today.strftime('%A'),
        'today': today,
        'total_employees': total_employees,
        'present_today': present_today,
        'absent_today': total_employees - present_today,
        'late_today_count': late_today_count,
        'monthly_penalty_count': monthly_penalty_count,
        'total_monthly_penalty': total_monthly_penalty,
        'total_today_penalty': total_today_penalty,
    }
    return render(request, 'employee_list.html', context)


@login_required
@admin_or_hr_required
def add_employee(request):
    """Yangi xodim qo'shish - admin yoki HR uchun"""
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            face_image_data = request.POST.get('face_image')
            if face_image_data:
                try:
                    format, imgstr = face_image_data.split(';base64,')
                    ext = format.split('/')[-1]
                    import uuid
                    filename = f"employee_{uuid.uuid4().hex}.{ext}"
                    from django.core.files.base import ContentFile
                    import base64
                    employee.photo.save(filename, ContentFile(base64.b64decode(imgstr)), save=False)
                except Exception:
                    pass
            employee.save()
            messages.success(request, f'{employee.full_name} muvaffaqiyatli qoshildi!')
            return redirect('employee_list')
    else:
        form = EmployeeForm()

    days = [
        {'code': 'monday', 'name': 'Dushanba'},
        {'code': 'tuesday', 'name': 'Seshanba'},
        {'code': 'wednesday', 'name': 'Chorshanba'},
        {'code': 'thursday', 'name': 'Payshanba'},
        {'code': 'friday', 'name': 'Juma'},
        {'code': 'saturday', 'name': 'Shanba'},
        {'code': 'sunday', 'name': 'Yakshanba'},
    ]

    return render(request, 'employee_form_face.html', {
        'form': form,
        'title': 'Yangi xodim qoshish',
        'days': days,
    })


@login_required
@admin_or_hr_required
def edit_employee(request, id):
    """Xodimni tahrirlash - admin yoki HR uchun"""
    employee = get_object_or_404(Employee, id=id)

    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            employee = form.save(commit=False)
            face_image_data = request.POST.get('face_image')
            if face_image_data:
                try:
                    format, imgstr = face_image_data.split(';base64,')
                    ext = format.split('/')[-1]
                    import uuid
                    filename = f"employee_{uuid.uuid4().hex}.{ext}"
                    from django.core.files.base import ContentFile
                    import base64
                    employee.photo.save(filename, ContentFile(base64.b64decode(imgstr)), save=False)
                except Exception:
                    pass
            employee.save()
            messages.success(request, f'{employee.full_name} muvaffaqiyatli yangilandi!')
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)

    days = [
        {'code': 'monday', 'name': 'Dushanba'},
        {'code': 'tuesday', 'name': 'Seshanba'},
        {'code': 'wednesday', 'name': 'Chorshanba'},
        {'code': 'thursday', 'name': 'Payshanba'},
        {'code': 'friday', 'name': 'Juma'},
        {'code': 'saturday', 'name': 'Shanba'},
        {'code': 'sunday', 'name': 'Yakshanba'},
    ]

    return render(request, 'employee_form_face.html', {
        'form': form,
        'title': 'Xodimni tahrirlash',
        'employee': employee,
        'days': days,
    })


@login_required
@admin_or_hr_required
def employee_detail(request, id):
    """Xodim ma'lumotlari - admin yoki HR uchun"""
    employee = get_object_or_404(Employee, id=id)
    attendances = Attendance.objects.filter(employee=employee).order_by('-date', '-time')[:50]
    salaries = MonthlySalary.objects.filter(employee=employee).order_by('-year', '-month')[:12]
    today_schedule = employee.get_today_schedule()

    context = {
        'employee': employee,
        'attendances': attendances,
        'salaries': salaries,
        'today_schedule': today_schedule,
        'work_days': employee.work_days,
        'work_schedule': employee.work_schedule,
    }
    return render(request, 'employee_detail.html', context)

@csrf_exempt
@login_required
@admin_or_hr_required
def delete_employee(request, id):
    """Xodimni o'chirish - admin yoki HR uchun"""
    print(f"DEBUG: O'chirish so'rovi kelgan. ID: {id}")  # DEBUG uchun

    if request.method == 'POST':
        try:
            # Xodimni topish
            employee = Employee.objects.get(id=id)
            employee_name = employee.full_name

            print(f"DEBUG: Xodim topildi: {employee_name}")  # DEBUG

            # 1. Avval Attendance (davomat) ma'lumotlarini o'chirish
            Attendance.objects.filter(employee=employee).delete()
            print("DEBUG: Attendance ma'lumotlari o'chirildi")  # DEBUG

            # 2. MonthlySalary ma'lumotlarini o'chirish
            MonthlySalary.objects.filter(employee=employee).delete()
            print("DEBUG: MonthlySalary ma'lumotlari o'chirildi")  # DEBUG

            # 3. Xodimni o'zini o'chirish
            employee.delete()
            print(f"DEBUG: Xodim o'chirildi: {employee_name}")  # DEBUG

            # Muvaffaqiyatli xabar
            messages.success(request, f'{employee_name} xodimi va uning barcha ma\'lumotlari o\'chirildi!')

            # AJAX so'rovi bo'lsa
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': f'{employee_name} o\'chirildi!',
                    'redirect': '/employees/'
                })
            else:
                # Oddiy so'rov bo'lsa
                return redirect('employee_list')

        except Employee.DoesNotExist:
            error_msg = f"Xodim topilmadi (ID: {id})"
            print(f"DEBUG: {error_msg}")  # DEBUG

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': error_msg
                })
            else:
                messages.error(request, error_msg)
                return redirect('employee_list')

        except Exception as e:
            error_msg = f"Xodimni o'chirishda xatolik: {str(e)}"
            print(f"DEBUG: {error_msg}")  # DEBUG

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': error_msg
                })
            else:
                messages.error(request, error_msg)
                return redirect('employee_list')

    # Agar POST emas bo'lsa
    return JsonResponse({'status': 'error', 'message': 'Faqat POST so\'rovi qabul qilinadi'})



@login_required
def checkin_page(request):
    """Kelish sahifasi - barcha foydalanuvchilar uchun"""
    employees = Employee.objects.filter(is_active=True)
    today = date.today()

    today_checkins = Attendance.objects.filter(date=today, type='in')
    checked_in_today = list(today_checkins.values_list('employee_id', flat=True))

    today_code = today.strftime('%A').lower()

    employees_with_schedule = []
    for emp in employees:
        schedule = emp.get_daily_schedule(today_code)
        employees_with_schedule.append({
            'id': emp.id,
            'full_name': emp.full_name,
            'position': emp.position,
            'photo': emp.photo,
            'monthly_salary': emp.monthly_salary,
            'late_penalty_per_minute': emp.late_penalty_per_minute,
            'allowed_late_minutes': emp.allowed_late_minutes,
            'today_schedule': schedule,
            'is_work_day': schedule['is_work_day']
        })

    locations = Location.objects.filter(is_active=True)
    locations_json = json.dumps([{
        'name': loc.name,
        'lat': float(loc.latitude) if loc.latitude else 0,
        'lng': float(loc.longitude) if loc.longitude else 0,
        'radius': loc.radius_meters,
    } for loc in locations])

    context = {
        'employees': employees,
        'employees_with_schedule': employees_with_schedule,
        'today_checkins': today_checkins,
        'checked_in_today': checked_in_today,
        'attendance_type': 'in',
        'page_title': 'Kelish',
        'btn_color': 'success',
        'icon': 'sign-in-alt',
        'today_day': today.strftime('%A'),
        'locations_json': locations_json,
    }
    return render(request, 'attendance_face.html', context)


@login_required
def checkout_page(request):
    """Chiqish sahifasi - barcha foydalanuvchilar uchun"""
    employees = Employee.objects.filter(is_active=True)
    today = date.today()

    today_checkins = Attendance.objects.filter(date=today, type='in')
    today_checkouts = Attendance.objects.filter(date=today, type='out')

    checked_out_today = list(today_checkouts.values_list('employee_id', flat=True))

    employees_with_checkin = []
    for emp in employees:
        if today_checkins.filter(employee=emp).exists() and not today_checkouts.filter(employee=emp).exists():
            employees_with_checkin.append(emp)

    locations = Location.objects.filter(is_active=True)
    locations_json = json.dumps([{
        'name': loc.name,
        'lat': float(loc.latitude) if loc.latitude else 0,
        'lng': float(loc.longitude) if loc.longitude else 0,
        'radius': loc.radius_meters,
    } for loc in locations])

    context = {
        'employees': employees_with_checkin,
        'today_checkouts': today_checkouts,
        'checked_out_today': checked_out_today,
        'attendance_type': 'out',
        'page_title': 'Chiqish',
        'btn_color': 'danger',
        'icon': 'sign-out-alt',
        'locations_json': locations_json,
    }
    return render(request, 'attendance_face.html', context)


# ============ DAVOMAT BELGILASH API (barcha foydalanuvchilar) ============

@csrf_exempt
@login_required
def mark_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            employee_id = data.get('employee_id')
            attendance_type = data.get('type', 'in')
            image_data = data.get('image_data', '')  # Base64 image from camera
            building_name = data.get('building', '')
            latitude = data.get('latitude')
            longitude = data.get('longitude')

            employee = get_object_or_404(Employee, id=employee_id)

            # Find location by building name or coordinates
            location = None
            if building_name and building_name != 'Noma\'lum':
                location = Location.objects.filter(name__iexact=building_name, is_active=True).first()
            if not location and latitude and longitude:
                try:
                    lat = float(latitude); lng = float(longitude)
                    for loc in Location.objects.filter(is_active=True, latitude__isnull=False, longitude__isnull=False):
                        from math import radians, sin, cos, sqrt, atan2
                        dlat = radians(lat - float(loc.latitude))
                        dlng = radians(lng - float(loc.longitude))
                        a = sin(dlat/2)**2 + cos(radians(float(loc.latitude))) * cos(radians(lat)) * sin(dlng/2)**2
                        dist = 6371000 * 2 * atan2(sqrt(a), sqrt(1-a))
                        if dist <= (loc.radius_meters or 100):
                            location = loc; break
                except Exception:
                    pass
            today = date.today()
            now_time = datetime.now().time()

            today_attendance = Attendance.objects.filter(
                employee=employee,
                date=today,
                type=attendance_type
            )

            if today_attendance.exists():
                last_record = today_attendance.latest('time')
                return JsonResponse({
                    'status': 'error',
                    'message': f'{employee.full_name} {last_record.type_display.lower()}i allaqachon qayd etilgan ({last_record.time})'
                })

            if attendance_type == 'out':
                checkin_exists = Attendance.objects.filter(
                    employee=employee,
                    date=today,
                    type='in'
                ).exists()

                if not checkin_exists:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'{employee.full_name} hali kelmagan. Avval kelishini belgilang!'
                    })

            late_status = 'ontime'
            late_minutes = 0
            penalty_amount = 0
            is_work_day = True

            if attendance_type == 'in':
                schedule = employee.get_today_schedule()
                is_work_day = schedule['is_work_day']

                if not is_work_day:
                    late_status = 'day_off'
                    message = f"{employee.full_name} kelishi qayd etildi. (Bugun dam olish kuni - ish kuni hisoblanmaydi)"
                else:
                    penalty_amount = employee.check_late_penalty(now_time, today)
                    if penalty_amount > 0:
                        late_status = 'late'
                        scheduled_time = datetime.strptime(schedule['start'], '%H:%M').time()
                        late_minutes_total = (datetime.combine(today, now_time) - datetime.combine(today, scheduled_time)).seconds // 60
                        late_minutes = max(0, late_minutes_total - employee.allowed_late_minutes)
                        message = f"{employee.full_name} kelishi qayd etildi. ({late_minutes} daqiqa kechikdi - Jarima: {penalty_amount:,.0f} so'm)"
                    else:
                        message = f"{employee.full_name} kelishi qayd etildi."
            else:
                message = f"{employee.full_name} chiqishi qayd etildi."

            attendance = Attendance.objects.create(
                employee=employee,
                date=today,
                time=now_time,
                type=attendance_type,
                status=late_status,
                late_minutes=late_minutes,
                penalty_amount=penalty_amount,
                notes=''
            )

            # Save face photo if provided
            if image_data:
                try:
                    format, imgstr = image_data.split(';base64,')
                    ext = format.split('/')[-1]
                    filename = f"face_{employee_id}_{attendance_type}_{today.strftime('%Y%m%d')}_{uuid.uuid4().hex}.{ext}"
                    photo_file = ContentFile(base64.b64decode(imgstr))

                    attendance.face_photo.save(filename, photo_file, save=True)

                    FaceCapture.objects.create(
                        employee=employee,
                        attendance=attendance,
                        photo=attendance.face_photo,
                        attendance_type=attendance_type,
                        location=location,
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                except Exception:
                    pass  # Photo save failure should not block attendance

            # Create LateAbsenceRecord if late
            if attendance_type == 'in' and late_status == 'late':
                LateAbsenceRecord.objects.get_or_create(
                    employee=employee,
                    date=today,
                    record_type='late',
                    defaults={
                        'late_minutes': late_minutes,
                        'attendance': attendance
                    }
                )

            # Send Telegram notification
            try:
                send_telegram_notification(attendance)
            except Exception:
                pass

            response_data = {
                'status': 'success',
                'message': message,
                'employee_name': employee.full_name,
                'time': attendance.time.strftime('%H:%M'),
                'type': attendance.type,
                'type_display': attendance.type_display,
                'late_status': late_status,
                'is_work_day': is_work_day,
                'penalty_amount': float(penalty_amount),
                'late_minutes': late_minutes
            }

            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Faqat POST soʻrovi qabul qilinadi'})


# ============ HISOBOTLAR (admin yoki HR) ============

@login_required
@admin_or_hr_required
def attendance_report(request):
    """Davomat hisoboti - admin yoki HR uchun"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_id = request.GET.get('employee_id')
    status = request.GET.get('status')

    attendances = Attendance.objects.all().order_by('-date', '-time')

    if start_date:
        attendances = attendances.filter(date__gte=start_date)
    if end_date:
        attendances = attendances.filter(date__lte=end_date)
    if employee_id:
        attendances = attendances.filter(employee_id=employee_id)
    if status:
        attendances = attendances.filter(status=status)

    paginator = Paginator(attendances, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    employees = Employee.objects.filter(is_active=True)

    context = {
        'attendances': page_obj,
        'employees': employees,
        'start_date': start_date,
        'end_date': end_date,
        'selected_employee': employee_id,
        'selected_status': status,
        'total_count': attendances.count(),
        'ontime_count': attendances.filter(status='ontime').count(),
        'late_count': attendances.filter(status='late').count(),
        'early_count': attendances.filter(status='early').count(),
        'day_off_count': attendances.filter(status='day_off').count(),
    }
    return render(request, 'attendance_report.html', context)


@login_required
@admin_or_hr_required
def daily_attendance_report(request):
    """Kunlik hisobot - admin yoki HR uchun"""
    report_date = request.GET.get('date')
    if report_date:
        report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
    else:
        report_date = date.today()

    employees = Employee.objects.filter(is_active=True)
    total_employees = employees.count()

    attendances = Attendance.objects.filter(date=report_date)
    checkins = attendances.filter(type='in')
    checkouts = attendances.filter(type='out')
    checkout_by_employee = {c.employee_id: c for c in checkouts}

    # Faqat ish kunlaridagi kelishlarni hisoblaymiz
    present_employees = employees.filter(
        attendances__date=report_date,
        attendances__type='in',
        attendances__status__in=['ontime', 'late', 'early']
    ).distinct()

    # Ish kunlarida kelmaganlar
    absent_employees = []
    for emp in employees:
        day_code = report_date.strftime('%A').lower()
        schedule = emp.get_daily_schedule(day_code)

        if schedule['is_work_day']:
            if not present_employees.filter(id=emp.id).exists():
                absent_employees.append(emp)

    late_count = checkins.filter(status='late').count()
    early_count = checkins.filter(status='early').count()
    day_off_count = checkins.filter(status='day_off').count()
    present_count = present_employees.count()

    # Combine checkins with their matching checkout
    combined_attendances = []
    for cin in checkins:
        cout = checkout_by_employee.get(cin.employee_id)
        combined_attendances.append({
            'checkin': cin,
            'checkout': cout.time if cout else None,
        })

    context = {
        'report_date': report_date,
        'employees': employees,
        'total_employees': total_employees,
        'checkins': checkins,
        'combined_attendances': combined_attendances,
        'checkouts': checkouts,
        'present_employees': present_employees,
        'absent_employees': absent_employees,
        'present_count': present_count,
        'late_count': late_count,
        'early_count': early_count,
        'absent_count': len(absent_employees),
        'day_off_count': day_off_count,
    }

    return render(request, 'daily_report.html', context)


@login_required
@admin_or_hr_required
def weekly_attendance_report(request):
    """Haftalik hisobot - admin yoki HR uchun"""
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    start_date = request.GET.get('start_date', start_of_week.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', end_of_week.strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date = start_of_week
        end_date = end_of_week

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    all_attendances = Attendance.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date', 'time')

    employees = Employee.objects.filter(is_active=True)
    weekly_stats = []

    week_days = []
    current = start_date
    while current <= end_date:
        week_days.append({
            'date': current
        })
        current += timedelta(days=1)

    for employee in employees:
        employee_attendances = all_attendances.filter(employee=employee)

        employee_checkins = employee_attendances.filter(type='in')
        employee_checkouts = employee_attendances.filter(type='out')

        present_days = employee_checkins.count()

        days_data = {}
        current_date = start_date

        while current_date <= end_date:
            day_key = current_date.isoformat()
            day_checkin = employee_checkins.filter(date=current_date).first()

            if day_checkin:
                day_checkout = employee_checkouts.filter(date=current_date).first()

                late_hours = day_checkin.late_minutes // 60
                late_minutes = day_checkin.late_minutes % 60

                days_data[day_key] = {
                    'checkin': day_checkin.time,
                    'checkout': day_checkout.time if day_checkout else None,
                    'checkin_display': day_checkin.date.strftime('%d.%m.%Y') + ' ' + day_checkin.time.strftime('%H:%M'),
                    'checkout_display': (day_checkout.date.strftime('%d.%m.%Y') + ' ' + day_checkout.time.strftime('%H:%M')) if day_checkout else None,
                    'status': day_checkin.status,
                    'late_minutes': day_checkin.late_minutes,
                    'late_hours': late_hours,
                    'late_minutes_remainder': late_minutes,
                }
            else:
                days_data[day_key] = None

            current_date += timedelta(days=1)

        late_days = employee_checkins.filter(status='late').count()
        weekly_stats.append({
            'employee': employee,
            'present_days': present_days,
            'days_data': days_data,
            'late_days': late_days,
            'early_days': employee_checkins.filter(status='early').count(),
            'early_leave_days': 0,
        })

    total_present = sum(s['present_days'] for s in weekly_stats)
    total_late = sum(s['late_days'] for s in weekly_stats)

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'weekly_stats': weekly_stats,
        'total_employees': employees.count(),
        'week_days': week_days,
        'total_present': total_present,
        'total_late': total_late,
        'total_absent': 0,
    }
    return render(request, 'weekly_report.html', context)

@login_required
@admin_or_hr_required
def monthly_attendance_report(request):
    """Oylik hisobot - admin yoki HR uchun"""
    today = date.today()
    start_of_month = date(today.year, today.month, 1)

    if today.month == 12:
        end_of_month = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(today.year, today.month + 1, 1) - timedelta(days=1)

    month = request.GET.get('month', today.month)
    year = request.GET.get('year', today.year)

    try:
        month = int(month)
        year = int(year)
        start_of_month = date(year, month, 1)
        if month == 12:
            end_of_month = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_of_month = date(year, month + 1, 1) - timedelta(days=1)
    except ValueError:
        month = today.month
        year = today.year

    # **ASOSIY O'ZGARISH: Faqat ish kunlaridagi kelishlarni hisoblaymiz**
    attendances = Attendance.objects.filter(
        date__gte=start_of_month,
        date__lte=end_of_month,
        type='in'
    ).order_by('date', 'time')

    employees = Employee.objects.filter(is_active=True)
    monthly_stats = []

    # Umumiy statistikalar
    total_present = 0
    total_absent = 0
    total_late = 0
    total_dayoff = 0

    for employee in employees:
        # **ISH KUNLARI SONI**
        work_days_in_month = 0
        current_date = start_of_month
        while current_date <= end_of_month:
            if current_date.strftime('%A').lower() in employee.work_days:
                work_days_in_month += 1
            current_date += timedelta(days=1)

        # **FAQT ISH KUNLARIDAGI KELISHLARNI HISOBLASH**
        employee_attendances = attendances.filter(employee=employee)
        present_days = 0
        late_days = 0
        early_days = 0
        day_off_days = 0

        # Har bir kelishni alohida tekshiramiz
        for attendance in employee_attendances:
            day_code = attendance.date.strftime('%A').lower()

            # Agar ish kuni bo'lsa
            if day_code in employee.work_days:
                if attendance.status == 'day_off':
                    day_off_days += 1
                else:
                    present_days += 1
                    if attendance.status == 'late':
                        late_days += 1
                    elif attendance.status == 'early':
                        early_days += 1
            else:
                # Dam olish kuni bo'lsa, hech qanday hisobga olinmaydi
                pass

        # **ABSENT DAYS: Ish kunlari - (kelgan kunlar + day_off kunlar)**
        absent_days = max(0, work_days_in_month - (present_days + day_off_days))

        # **ATTENDANCE RATE: Faqat kelgan kunlar / ish kunlari**
        attendance_rate = 0
        if work_days_in_month > 0:
            attendance_rate = (present_days / work_days_in_month * 100)

        # Umumiy statistikaga qo'shish
        total_present += present_days
        total_absent += absent_days
        total_late += late_days
        total_dayoff += day_off_days

        monthly_stats.append({
            'employee': employee,
            'present_days': present_days,  # Faqat ish kunlaridagi kelishlar
            'work_days': work_days_in_month,
            'absent_days': absent_days,
            'late_days': late_days,
            'early_days': early_days,
            'day_off_days': day_off_days,
            'attendance_rate': attendance_rate,
        })

    months_list = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr'),
    ]

    current_year = date.today().year
    years = range(current_year - 5, current_year + 1)

    context = {
        'month': month,
        'year': year,
        'start_date': start_of_month,
        'end_date': end_of_month,
        'monthly_stats': monthly_stats,
        'months': months_list,
        'years': years,
        'total_employees': employees.count(),
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_dayoff': total_dayoff,
    }
    return render(request, 'monthly_report.html', context)

@login_required
@admin_or_hr_required
def export_monthly_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    start_of_month = date(year, month, 1)
    if month == 12:
        end_of_month = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(year, month + 1, 1) - timedelta(days=1)

    employees = Employee.objects.filter(is_active=True)
    attendances = Attendance.objects.filter(
        date__gte=start_of_month, date__lte=end_of_month, type='in'
    ).order_by('date', 'time')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2001FF", end_color="2001FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    headers = [
        "T/r", "Xodim", "Lavozimi", "Bo'lim",
        "Ish kunlari", "Kelgan", "Kelmagan",
        "Kechikish", "Erta kelish", "Dam olish",
        "Foiz", "Jarima"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row_num = 2
    for i, emp in enumerate(employees, 1):
        work_days = 0
        current = start_of_month
        while current <= end_of_month:
            if current.strftime('%A').lower() in emp.work_days:
                work_days += 1
            current += timedelta(days=1)

        emp_att = attendances.filter(employee=emp)
        present = late = early = day_off = 0
        penalty = 0
        for a in emp_att:
            day_code = a.date.strftime('%A').lower()
            if day_code in emp.work_days:
                if a.status == 'day_off':
                    day_off += 1
                else:
                    present += 1
                    if a.status == 'late':
                        late += 1
                        penalty += float(a.penalty_amount)
                    elif a.status == 'early':
                        early += 1

        absent = max(0, work_days - (present + day_off))
        rate = round((present / work_days * 100) if work_days > 0 else 0, 1)

        vals = [
            i, emp.full_name, emp.position, emp.department,
            work_days, present, absent, late, early, day_off,
            rate, penalty
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            cell.alignment = center
        row_num += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    for c in range(5, 13):
        ws.column_dimensions[get_column_letter(c)].width = 12

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="davomat-{year}-{month:02d}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_or_hr_required
def salary_report(request):
    """Maosh hisoboti - admin yoki HR uchun"""
    today = date.today()
    selected_year = int(request.GET.get('year', today.year))
    selected_month = int(request.GET.get('month', today.month))
    selected_employee = request.GET.get('employee_id')

    # Yillar va oylar
    years = list(range(today.year - 5, today.year + 1))
    years.reverse()

    months = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
    ]

    # Xodimlar
    employees = Employee.objects.filter(is_active=True)

    if selected_employee:
        employees = employees.filter(id=selected_employee)

    salaries = []
    total_salary = 0
    total_penalty = 0
    paid_count = 0
    unpaid_count = 0
    day_off_count = 0
    total_work_days = 0
    total_present_days = 0
    total_absent_days = 0
    total_late_days = 0

    # Majburiy hisoblash
    force_calculate = request.GET.get('calculate') in ['true', 'force']

    for employee in employees:
        try:
            # Maosh yozuvini olish yoki yaratish
            salary, created = MonthlySalary.objects.get_or_create(
                employee=employee,
                year=selected_year,
                month=selected_month,
                defaults={
                    'basic_salary': employee.monthly_salary or 0,
                    'is_paid': False
                }
            )

            # Hisoblash
            if force_calculate or created:
                salary.calculate_salary()
                salary.save()
            elif salary.present_days == 0 and salary.work_days > 0:
                # Agar kelgan kunlar 0 bo'lsa, lekin ish kunlari bor bo'lsa, hisoblaymiz
                salary.calculate_salary()
                salary.save()

            salaries.append(salary)

            # Umumiy statistikalar
            total_salary += salary.net_salary or 0
            total_penalty += salary.total_penalty or 0
            total_work_days += salary.work_days or 0
            total_present_days += salary.present_days or 0
            total_absent_days += salary.absent_days or 0
            total_late_days += salary.late_days or 0
            day_off_count += salary.day_off_days or 0

            if salary.is_paid:
                paid_count += 1
            else:
                unpaid_count += 1

        except Exception as e:
            print(f"Xato: {employee.full_name} uchun maosh hisoblashda xatolik: {e}")
            continue

    # Kontekst
    context = {
        'salaries': salaries,
        'employees': Employee.objects.filter(is_active=True),
        'years': years,
        'months': months,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_employee': int(selected_employee) if selected_employee else '',
        'total_salary': total_salary,
        'total_penalty': total_penalty,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'day_off_count': day_off_count,
        'total_work_days': total_work_days,
        'total_present_days': total_present_days,
        'total_absent_days': total_absent_days,
        'total_late_days': total_late_days,
    }

    return render(request, 'salary_report.html', context)


@login_required
@admin_or_hr_required
def calculate_monthly_salary(request, employee_id, year, month):
    """Maosh hisoblash - admin yoki HR uchun"""
    employee = get_object_or_404(Employee, id=employee_id)

    salary, created = MonthlySalary.objects.get_or_create(
        employee=employee,
        year=year,
        month=month,
        defaults={'basic_salary': employee.monthly_salary}
    )

    net_salary = salary.calculate_salary()

    return JsonResponse({
        'status': 'success',
        'message': f'{employee.full_name} uchun {month}/{year} oyi maoshi hisoblandi',
        'net_salary': float(net_salary),
        'basic_salary': float(salary.basic_salary),
        'total_penalty': float(salary.total_penalty),
        'present_days': salary.present_days,
        'late_days': salary.late_days,
        'absent_days': salary.absent_days,
        'day_off_days': salary.day_off_days,
    })


@csrf_exempt
@login_required
@admin_or_hr_required
def mark_salary_paid(request, salary_id):
    """Maosh to'landi deb belgilash - admin yoki HR uchun"""
    salary = get_object_or_404(MonthlySalary, id=salary_id)

    if request.method == 'POST':
        salary.is_paid = True
        salary.paid_date = date.today()
        salary.save()

        messages.success(request, f'{salary.employee.full_name} uchun {salary.month}/{salary.year} oyi maoshi to\'landi deb belgilandi')
        redirect_url = reverse('salary_report')
        params = {}
        year = request.POST.get('year')
        month = request.POST.get('month')
        emp_id = request.POST.get('employee_id')
        if year: params['year'] = year
        if month: params['month'] = month
        if emp_id: params['employee_id'] = emp_id
        if params:
            redirect_url += '?' + '&'.join(f'{k}={v}' for k, v in params.items())
        return redirect(redirect_url)

    return JsonResponse({'status': 'error', 'message': 'Faqat POST soʻrovi'})


# ============ GALLERY DELETE PHOTO ============

@login_required
@admin_or_hr_required
def gallery_delete_photo(request, record_id):
    """Galereyadagi rasmni va unga bog'liq davomatni o'chirish"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST soʻrovi'})

    source = request.POST.get('source', '')
    delete_attendance = request.POST.get('delete_attendance') == 'true'

    try:
        if source == 'face_capture':
            obj = get_object_or_404(FaceCapture, id=record_id)
            attendance = obj.attendance
            # Delete photo file
            if obj.photo:
                storage = obj.photo.storage
                if storage.exists(obj.photo.name):
                    storage.delete(obj.photo.name)
            obj.delete()
            message = 'Rasm o\'chirildi'
            if delete_attendance and attendance:
                attendance.delete()
                message += ' va davomat o\'chirildi'
        elif source == 'attendance':
            obj = get_object_or_404(Attendance, id=record_id)
            if obj.face_photo:
                storage = obj.face_photo.storage
                if storage.exists(obj.face_photo.name):
                    storage.delete(obj.face_photo.name)
            if delete_attendance:
                obj.delete()
                message = 'Rasm va davomat o\'chirildi'
            else:
                obj.face_photo = None
                obj.save()
                message = 'Rasm o\'chirildi (davomat saqlandi)'
        else:
            return JsonResponse({'status': 'error', 'message': 'Noto\'g\'ri manba'})

        return JsonResponse({'status': 'success', 'message': message})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ============ ADMIN CLEANUP SETTINGS ============

@login_required
@admin_or_hr_required
def cleanup_settings(request):
    """Eski ma'lumotlarni ko'rish va tozalash sahifasi"""
    today = date.today()

    if request.method == 'POST' and request.POST.get('action') == 'save_auto':
        config, _ = CleanupConfig.objects.get_or_create(pk=1)
        config.auto_enabled = request.POST.get('auto_enabled') == 'on'
        config.months_back = int(request.POST.get('months_back', 2))
        config.interval_days = int(request.POST.get('interval_days', 60))
        config.clean_face_captures = request.POST.get('clean_face_captures') == 'on'
        config.clean_attendance_photos = request.POST.get('clean_attendance_photos') == 'on'
        config.clean_attendance_records = request.POST.get('clean_attendance_records') == 'on'
        config.save()
        messages.success(request, 'Avtomatik tozalash sozlamalari saqlandi')
        return redirect('cleanup_settings')

    target_year = request.GET.get('target_year')
    target_month = request.GET.get('target_month')
    months_back = int(request.GET.get('months', 2))
    cutoff_date = today - timedelta(days=months_back * 30)

    if target_year and target_month:
        y, m = int(target_year), int(target_month)
        month_start = date(y, m, 1)
        if m == 12:
            month_end = date(y + 1, 1, 1)
        else:
            month_end = date(y, m + 1, 1)

        old_face_captures = FaceCapture.objects.filter(captured_at__date__gte=month_start, captured_at__date__lt=month_end).order_by('-captured_at')[:100]
        old_attendance_photos = Attendance.objects.exclude(face_photo='').filter(date__gte=month_start, date__lt=month_end).order_by('-date')[:100]
        old_attendance_records = Attendance.objects.filter(date__gte=month_start, date__lt=month_end).order_by('-date')[:100]

        total_face_captures = FaceCapture.objects.filter(captured_at__date__gte=month_start, captured_at__date__lt=month_end).count()
        total_attendance_photos = Attendance.objects.exclude(face_photo='').filter(date__gte=month_start, date__lt=month_end).count()
        total_attendance_records = Attendance.objects.filter(date__gte=month_start, date__lt=month_end).count()
    else:
        old_face_captures = FaceCapture.objects.filter(captured_at__date__lt=cutoff_date).order_by('-captured_at')[:100]
        old_attendance_photos = Attendance.objects.exclude(face_photo='').filter(date__lt=cutoff_date).order_by('-date')[:100]
        old_attendance_records = Attendance.objects.filter(date__lt=cutoff_date).order_by('-date')[:100]

        total_face_captures = FaceCapture.objects.filter(captured_at__date__lt=cutoff_date).count()
        total_attendance_photos = Attendance.objects.exclude(face_photo='').filter(date__lt=cutoff_date).count()
        total_attendance_records = Attendance.objects.filter(date__lt=cutoff_date).count()

    auto_config = CleanupConfig.objects.filter(pk=1).first()

    months_list = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
    ]
    years_list = list(range(today.year - 3, today.year + 1))
    years_list.reverse()

    context = {
        'target_year': int(target_year) if target_year else None,
        'target_month': int(target_month) if target_month else None,
        'month_start': month_start if target_year and target_month else None,
        'month_end': month_end if target_year and target_month else None,
        'months_back': months_back,
        'cutoff_date': cutoff_date,
        'old_face_captures': old_face_captures,
        'old_attendance_photos': old_attendance_photos,
        'old_attendance_records': old_attendance_records,
        'total_face_captures': total_face_captures,
        'total_attendance_photos': total_attendance_photos,
        'total_attendance_records': total_attendance_records,
        'auto_config': auto_config,
        'months_list': months_list,
        'years_list': years_list,
    }
    return render(request, 'cleanup_settings.html', context)


@login_required
@admin_or_hr_required
def run_cleanup(request):
    """Tozalashni amalga oshirish (oylik yoki eski ma'lumotlar)"""
    if request.method != 'POST':
        messages.error(request, 'Faqat POST soʻrovi')
        return redirect('cleanup_settings')

    target_year = request.POST.get('target_year')
    target_month = request.POST.get('target_month')
    data_types = request.POST.getlist('data_types')

    try:
        if target_year and target_month:
            y, m = int(target_year), int(target_month)
            month_start = date(y, m, 1)
            if m == 12:
                month_end = date(y + 1, 1, 1)
            else:
                month_end = date(y, m + 1, 1)
        else:
            months_back = int(request.POST.get('months', 2))
            cutoff_date = date.today() - timedelta(days=months_back * 30)

        deleted_parts = []
        if 'face_captures' in data_types:
            if target_year and target_month:
                qs = FaceCapture.objects.filter(captured_at__date__gte=month_start, captured_at__date__lt=month_end)
            else:
                qs = FaceCapture.objects.filter(captured_at__date__lt=cutoff_date)
            count = qs.count()
            for obj in qs:
                if obj.photo:
                    try:
                        storage = obj.photo.storage
                        if storage.exists(obj.photo.name):
                            storage.delete(obj.photo.name)
                    except Exception:
                        pass
            qs.delete()
            if count:
                deleted_parts.append(f"Face ID rasmlar: {count}")

        if 'attendance_photos' in data_types:
            if target_year and target_month:
                qs = Attendance.objects.exclude(face_photo='').filter(date__gte=month_start, date__lt=month_end)
            else:
                qs = Attendance.objects.exclude(face_photo='').filter(date__lt=cutoff_date)
            count = qs.count()
            for obj in qs:
                if obj.face_photo:
                    try:
                        storage = obj.face_photo.storage
                        if storage.exists(obj.face_photo.name):
                            storage.delete(obj.face_photo.name)
                    except Exception:
                        pass
            qs.delete()
            if count:
                deleted_parts.append(f"Davomat rasmlar: {count}")

        if 'attendance_records' in data_types:
            if target_year and target_month:
                qs = Attendance.objects.filter(date__gte=month_start, date__lt=month_end)
            else:
                qs = Attendance.objects.filter(date__lt=cutoff_date)
            count = qs.count()
            qs.delete()
            if count:
                deleted_parts.append(f"Davomat yozuvlari: {count}")

        if deleted_parts:
            messages.success(request, 'Tozalash yakunlandi: ' + ', '.join(deleted_parts))
        else:
            messages.info(request, 'O\'chiriladigan ma\'lumot topilmadi')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')

    return redirect('cleanup_settings')


# ============ QO'SHIMCHA API (barcha foydalanuvchilar) ============

@login_required
def get_employee_schedule(request, employee_id):
    """Xodim jadvali - barcha foydalanuvchilar uchun"""
    employee = get_object_or_404(Employee, id=employee_id)

    today_schedule = employee.get_today_schedule()
    weekly_schedule = {}

    for day_code, day_name in Employee.WORK_DAYS_CHOICES:
        schedule = employee.get_daily_schedule(day_code)
        weekly_schedule[day_code] = {
            'name': day_name,
            'start': schedule['start'],
            'end': schedule['end'],
            'is_work_day': schedule['is_work_day'],
        }

    return JsonResponse({
        'status': 'success',
        'employee_name': employee.full_name,
        'today_schedule': today_schedule,
        'weekly_schedule': weekly_schedule,
    })


@login_required
def get_today_attendance(request):
    """Bugungi davomat - barcha foydalanuvchilar uchun"""
    today = date.today()

    checkins = Attendance.objects.filter(date=today, type='in')
    checkouts = Attendance.objects.filter(date=today, type='out')

    checkins_list = [{
        'id': att.id,
        'employee_name': att.employee.full_name,
        'time': att.time.strftime('%H:%M'),
        'status': att.status,
        'late_minutes': att.late_minutes,
        'penalty_amount': float(att.penalty_amount),
    } for att in checkins]

    checkouts_list = [{
        'id': att.id,
        'employee_name': att.employee.full_name,
        'time': att.time.strftime('%H:%M'),
    } for att in checkouts]

    return JsonResponse({
        'status': 'success',
        'date': today.strftime('%Y-%m-%d'),
        'checkins': checkins_list,
        'checkouts': checkouts_list,
        'total_checkins': len(checkins_list),
        'total_checkouts': len(checkouts_list),
    })


# ============ JARIMA TARIXI API ============

@login_required
def get_employee_penalty_history(request, employee_id):
    """Xodim jarima tarixi"""
    employee = get_object_or_404(Employee, id=employee_id)

    # Get penalty history for last 6 months
    penalty_history = employee.get_monthly_penalty_history(limit=6)

    return JsonResponse({
        'status': 'success',
        'employee_name': employee.full_name,
        'history': penalty_history
    })


# ============ OYLIK JARIMA HISOBOTI ============

@login_required
@admin_or_hr_required
def monthly_penalty_report(request):
    """Oylik jarimalar hisoboti"""
    today = date.today()

    # Filter by month
    selected_year = int(request.GET.get('year', today.year))
    selected_month = int(request.GET.get('month', today.month))
    selected_employee = request.GET.get('employee_id')

    # Get all employees
    employees = Employee.objects.filter(is_active=True)

    if selected_employee:
        employees = employees.filter(id=selected_employee)

    # Calculate penalties for each employee
    penalty_data = []
    total_penalty = 0
    total_late_days = 0
    total_late_minutes = 0

    for employee in employees:
        # Get penalty details for selected month
        penalty_details = employee.get_month_penalty_details(selected_year, selected_month)

        # Get daily penalties
        daily_penalties = penalty_details.get('daily_penalties', [])

        # Get penalty history for comparison
        penalty_history = employee.get_monthly_penalty_history(limit=3)

        penalty_data.append({
            'employee': employee,
            'penalty_details': penalty_details,
            'daily_penalties': daily_penalties,
            'penalty_history': penalty_history
        })

        # Update totals
        total_penalty += penalty_details['total_penalty']
        total_late_days += penalty_details['late_days']
        total_late_minutes += penalty_details['total_late_minutes']

    # Sort by penalty amount (descending)
    penalty_data.sort(key=lambda x: x['penalty_details']['total_penalty'], reverse=True)

    # Years and months for filter
    years = list(range(today.year - 2, today.year + 1))
    years.reverse()

    months = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
    ]

    context = {
        'penalty_data': penalty_data,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_employee': int(selected_employee) if selected_employee else '',
        'years': years,
        'months': months,
        'employees': Employee.objects.filter(is_active=True),
        'total_penalty': total_penalty,
        'total_late_days': total_late_days,
        'total_late_minutes': total_late_minutes,
        'today': today,
    }

    return render(request, 'monthly_penalty_report.html', context)


# ============ TELEGRAM BOT INTEGRATION ============

import requests

TELEGRAM_BOT_TOKEN = None  # Set in settings.py or environment
TELEGRAM_CHAT_ID = None    # Set in settings.py or environment

def get_telegram_config():
    from django.conf import settings
    token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None) or TELEGRAM_BOT_TOKEN
    chat_id = getattr(settings, 'TELEGRAM_CHAT_ID', None) or TELEGRAM_CHAT_ID
    return token, chat_id

def send_telegram_notification(attendance):
    token, chat_id = get_telegram_config()
    if not token or not chat_id:
        return False

    emp = attendance.employee
    type_text = "✅ KELDI" if attendance.type == 'in' else "❌ CHIQDI"
    status_text = dict(Attendance.STATUS_CHOICES).get(attendance.status, attendance.status)

    text = (
        f"<b>🆔 Face ID - {type_text}</b>\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"👤 <b>Xodim:</b> {emp.full_name}\n"
        f"💼 <b>Lavozim:</b> {emp.position}\n"
        f"🏢 <b>Bo'lim:</b> {emp.department}\n"
        f"📅 <b>Sana:</b> {attendance.date.strftime('%d.%m.%Y')}\n"
        f"⏰ <b>Vaqt:</b> {attendance.time.strftime('%H:%M')}\n"
        f"📊 <b>Holat:</b> {status_text}\n"
    )

    if attendance.late_minutes > 0:
        text += f"⏳ <b>Kechikish:</b> {attendance.late_minutes} daq.\n"
        text += f"💰 <b>Jarima:</b> {attendance.penalty_amount:,.0f} so'm\n"

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = requests.post(url, json={
            'chat_id': chat_id,
            'text': text,
            'parse_mode': 'HTML'
        }, timeout=10)
        return resp.ok
    except Exception:
        return False


# ============ FACE PHOTO GALLERY ============

@login_required
@admin_or_hr_required
def face_gallery(request):
    employee_id = request.GET.get('employee_id') or None
    selected_date = request.GET.get('date') or None
    selected_type = request.GET.get('type') or None

    face_captures = FaceCapture.objects.select_related('employee', 'location')
    attendance_photos = Attendance.objects.select_related('employee').exclude(face_photo='')

    employee_filter = {}
    date_filter_cap = {}
    date_filter_att = {}
    type_filter_cap = {}
    type_filter_att = {}

    if employee_id:
        employee_filter = {'employee_id': employee_id}
    if selected_date:
        date_filter_cap = {**date_filter_cap, 'captured_at__date': selected_date}
        date_filter_att = {**date_filter_att, 'date': selected_date}
    if selected_type:
        type_filter_cap = {**type_filter_cap, 'attendance_type': selected_type}
        type_filter_att = {**type_filter_att, 'type': selected_type}

    face_captures = face_captures.filter(**employee_filter, **date_filter_cap, **type_filter_cap)
    attendance_photos = attendance_photos.filter(**employee_filter, **date_filter_att, **type_filter_att)

    # Deduplicate: skip Attendance records that already have a matching FaceCapture
    face_capture_att_ids = set(fc.attendance_id for fc in face_captures if fc.attendance_id)
    face_capture_keys = set((fc.employee_id, fc.captured_at.date(), fc.attendance_type) for fc in face_captures)

    combined = []
    for fc in face_captures:
        loc = fc.location
        combined.append({
            'id': fc.id,
            'employee_name': fc.employee.full_name,
            'employee_id': fc.employee_id,
            'photo_url': fc.photo.url if fc.photo else '',
            'attendance_type': fc.attendance_type,
            'location_name': loc.name if loc else '',
            'location_address': loc.address if loc else '',
            'location_building': loc.building if loc else '',
            'latitude': str(loc.latitude) if loc and loc.latitude else '',
            'longitude': str(loc.longitude) if loc and loc.longitude else '',
            'captured_at': fc.captured_at,
            'source': 'face_capture',
        })
    for att in attendance_photos:
        if att.id in face_capture_att_ids:
            continue
        att_key = (att.employee_id, att.date, att.type)
        if att_key in face_capture_keys:
            continue
        naive_dt = datetime.combine(att.date, att.time)
        aware_dt = timezone.make_aware(naive_dt) if timezone.is_naive(naive_dt) else naive_dt
        combined.append({
            'id': att.id,
            'employee_name': att.employee.full_name,
            'employee_id': att.employee_id,
            'photo_url': att.face_photo.url if att.face_photo else '',
            'attendance_type': att.type,
            'location_name': '',
            'location_address': '',
            'location_building': '',
            'latitude': '',
            'longitude': '',
            'captured_at': aware_dt,
            'source': 'attendance',
        })

    combined.sort(key=lambda x: x['captured_at'], reverse=True)
    total_photos = len(combined)

    today_start_naive = datetime.combine(date.today(), time.min)
    today_start = timezone.make_aware(today_start_naive) if timezone.is_naive(today_start_naive) else today_start_naive
    today_photos = sum(1 for p in combined if p['captured_at'] >= today_start)

    employees = Employee.objects.filter(is_active=True).order_by('first_name')

    paginator = Paginator(combined, 24)
    page_number = request.GET.get('page')
    photos_page = paginator.get_page(page_number)

    context = {
        'photos': photos_page,
        'employees': employees,
        'selected_employee': employee_id,
        'selected_date': selected_date,
        'selected_type': selected_type,
        'total_photos': total_photos,
        'today_photos': today_photos,
    }
    return render(request, 'gallery.html', context)


# ============ LOCATION MANAGEMENT ============

@login_required
@admin_or_hr_required
def location_list(request):
    locations = Location.objects.all()
    locations_json = json.dumps([{
        'id': loc.id,
        'name': loc.name,
        'building': loc.building,
        'address': loc.address,
        'lat': str(loc.latitude) if loc.latitude else None,
        'lng': str(loc.longitude) if loc.longitude else None,
        'radius': loc.radius_meters,
        'branch_name': loc.branch_name,
        'is_active': loc.is_active,
    } for loc in locations])
    return render(request, 'location_list.html', {
        'locations': locations,
        'locations_json': locations_json,
    })


@login_required
@admin_or_hr_required
def location_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        building = request.POST.get('building', '').strip()
        address = request.POST.get('address', '').strip()
        latitude = request.POST.get('latitude', '').strip().replace(',', '.')
        longitude = request.POST.get('longitude', '').strip().replace(',', '.')
        radius_meters = request.POST.get('radius_meters', '100').strip()
        branch_name = request.POST.get('branch_name', '').strip()
        is_active = request.POST.get('is_active') == 'on'

        if not name or not building or not address:
            messages.error(request, 'Lokatsiya nomi, bino nomi va manzil maydonlari majburiy!')
            return render(request, 'location_form.html', {'location': None})

        try:
            Location.objects.create(
                name=name, building=building, address=address,
                latitude=latitude or None, longitude=longitude or None,
                radius_meters=int(radius_meters) if radius_meters else 100,
                branch_name=branch_name, is_active=is_active
            )
            messages.success(request, 'Lokatsiya qo\'shildi!')
            return redirect('location_list')
        except Exception as e:
            messages.error(request, f'Lokatsiya qo\'shishda xatolik: {str(e)}')
            return render(request, 'location_form.html', {'location': None})

    return render(request, 'location_form.html', {'location': None})


@login_required
@admin_or_hr_required
def location_edit(request, id):
    location = get_object_or_404(Location, id=id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        building = request.POST.get('building', '').strip()
        address = request.POST.get('address', '').strip()

        if not name or not building or not address:
            messages.error(request, 'Lokatsiya nomi, bino nomi va manzil maydonlari majburiy!')
            return render(request, 'location_form.html', {'location': location})

        try:
            location.name = name
            location.building = building
            location.address = address
            location.latitude = request.POST.get('latitude', '').strip().replace(',', '.') or None
            location.longitude = request.POST.get('longitude', '').strip().replace(',', '.') or None
            location.radius_meters = int(request.POST.get('radius_meters', '100').strip() or 100)
            location.branch_name = request.POST.get('branch_name', '').strip()
            location.is_active = request.POST.get('is_active') == 'on'
            location.save()
            messages.success(request, 'Lokatsiya yangilandi!')
            return redirect('location_list')
        except Exception as e:
            messages.error(request, f'Lokatsiyani tahrirlashda xatolik: {str(e)}')
            return render(request, 'location_form.html', {'location': location})

    return render(request, 'location_form.html', {'location': location})


@login_required
@admin_or_hr_required
def location_delete(request, id):
    if request.method == 'POST':
        location = get_object_or_404(Location, id=id)
        location.delete()
        messages.success(request, f'Lokatsiya o\'chirildi!')
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return redirect('location_list')
    return JsonResponse({'status': 'error', 'message': 'Faqat POST'}, status=405)


# ============ FORCE CACHE REFRESH ============

@login_required
@admin_only_required
def force_cache_refresh(request):
    if request.method == 'POST':
        employees = Employee.objects.all()
        for emp in employees:
            emp.cache_version += 1
            emp.save()
        messages.success(request, f'Barcha {employees.count()} xodimning kesh ma\'lumotlari yangilandi!')
        return redirect('dashboard')
    return redirect('dashboard')


# ============ LATE / ABSENCE MONITORING ============

@login_required
@admin_or_hr_required
def late_monitoring(request):
    record_type = request.GET.get('type', 'late')
    employee_id = request.GET.get('employee_id')

    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    records = LateAbsenceRecord.objects.select_related('employee', 'attendance').filter(
        record_type=record_type,
        date__gte=start_date,
        date__lt=end_date,
    )

    if employee_id:
        records = records.filter(employee_id=employee_id)

    total_minutes = records.aggregate(total=Sum('late_minutes'))['total'] or 0
    total_records = records.count()

    # Per-employee summary for the month
    emp_summary_raw = records.values('employee_id').annotate(
        total_count=Count('id'),
        total_late_minutes=Sum('late_minutes'),
    ).order_by('-total_count')
    employee_summary = []
    for item in emp_summary_raw:
        emp = Employee.objects.filter(id=item['employee_id']).first()
        if emp:
            employee_summary.append({
                'employee': emp,
                'total_count': item['total_count'],
                'total_late_minutes': item['total_late_minutes'] or 0,
            })

    paginator = Paginator(records, 50)
    page = request.GET.get('page')
    records_page = paginator.get_page(page)

    employees = Employee.objects.filter(is_active=True).order_by('first_name')

    month_name_uz = ['', 'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun', 'Iyul', 'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr']

    context = {
        'records': records_page,
        'employees': employees,
        'record_type': record_type,
        'selected_employee': employee_id,
        'month': month,
        'year': year,
        'month_name': month_name_uz[month],
        'start_date': start_date,
        'end_date': end_date,
        'total_records': total_records,
        'total_minutes': total_minutes,
        'total_employees': employees.count(),
        'employee_summary': employee_summary,
    }
    return render(request, 'late_monitoring.html', context)


@login_required
@admin_or_hr_required
def late_monitoring_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    record_type = request.GET.get('type', 'late')

    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    records = LateAbsenceRecord.objects.filter(
        employee=employee,
        record_type=record_type,
        date__gte=start_date,
        date__lt=end_date
    ).order_by('-date')

    months_list = [(1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
                   (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
                   (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')]
    current_year = date.today().year
    years = range(current_year - 5, current_year + 1)

    context = {
        'employee': employee,
        'records': records,
        'record_type': record_type,
        'month': month,
        'year': year,
        'months': months_list,
        'years': years,
    }
    return render(request, 'late_monitoring_detail.html', context)


@login_required
@admin_or_hr_required
def late_record_edit(request, record_id):
    record = get_object_or_404(LateAbsenceRecord, id=record_id)
    if request.method == 'POST':
        record.reason = request.POST.get('reason', '')
        record.admin_note = request.POST.get('admin_note', '')
        record.save()
        messages.success(request, 'Ma\'lumot yangilandi!')
        from_list = request.POST.get('from_list')
        if from_list:
            return redirect(f"{reverse('late_monitoring')}?type={record.record_type}")
        return redirect('late_monitoring_detail', employee_id=record.employee_id)
    return render(request, 'late_monitoring_form.html', {'record': record})


# ============ REPORTS - PDF EXPORT ============

@login_required
@admin_or_hr_required
def export_report_pdf(request):
    report_type = request.GET.get('type', 'daily')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    today = date.today()

    if report_type == 'daily':
        start_date = end_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today
    elif report_type == 'weekly':
        start_of_week = today - timedelta(days=today.weekday())
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else start_of_week
        end_date = start_date + timedelta(days=6)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else date(today.year, today.month, 1)
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    attendances = Attendance.objects.filter(
        date__gte=start_date, date__lte=end_date
    ).select_related('employee').order_by('date', 'time')

    employees_data = []
    for att in attendances:
        employees_data.append({
            'employee': att.employee,
            'date': att.date,
            'time': att.time,
            'type': att.type_display,
            'status': att.get_status_display(),
            'late_minutes': att.late_minutes,
            'penalty_amount': att.penalty_amount,
        })

    total_count = len(employees_data)
    ontime_count = attendances.filter(status='ontime').count()
    late_count = attendances.filter(status='late').count()
    absent_count = attendances.filter(status='absent').count()
    day_off_count = attendances.filter(status='day_off').count()

    context = {
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date,
        'employees_data': employees_data,
        'total_stats': {
            'total': total_count,
            'ontime': ontime_count,
            'late': late_count,
            'absent': absent_count,
            'day_off': day_off_count,
        }
    }
    return render(request, 'report_pdf.html', context)


# ============ EXPORT EXCEL (Enhanced) ============

@login_required
@admin_or_hr_required
def export_excel_report(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = date.today()

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else date(today.year, today.month, 1)
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except (ValueError, TypeError):
        start_date = date(today.year, today.month, 1)
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    attendances = Attendance.objects.filter(
        date__gte=start_date, date__lte=end_date
    ).select_related('employee').order_by('date', 'time')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['T/r', 'Xodim', 'Lavozimi', "Bo'lim", 'Sana', 'Vaqt', 'Tur', 'Holat', 'Kechikish (daq.)', 'Jarima']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, att in enumerate(attendances, 1):
        row = i + 1
        data = [
            i, att.employee.full_name, att.employee.position,
            att.employee.department, att.date.strftime('%d.%m.%Y'),
            att.time.strftime('%H:%M'), att.type_display,
            att.get_status_display(), att.late_minutes,
            float(att.penalty_amount)
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in [1, 5, 6, 7, 8, 9, 10] else 'left')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=hisobot_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response


# ============ FIX ACCOUNTS - PASSWORD CHANGE ============

@login_required
def password_change_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Parolingiz muvaffaqiyatli o\'zgartirildi!')
            return redirect('profile')
        else:
            messages.error(request, 'Iltimos, xatolarni to\'g\'rilang.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'admin/password_change_form.html', {'form': form})


# ============ OLD-STYLE EMPLOYEE FORM ============

@login_required
@admin_or_hr_required
def add_employee_old(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'{employee.full_name} muvaffaqiyatli qo\'shildi!')
            return redirect('employee_list')
        else:
            messages.error(request, 'Iltimos, xatolarni to\'g\'rilang.')
    else:
        form = EmployeeForm()
    return render(request, 'employee_form_old.html', {
        'form': form,
        'employee': None,
        'is_add': True,
    })


@login_required
@admin_or_hr_required
def edit_employee_old(request, id):
    employee = get_object_or_404(Employee, id=id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f'{employee.full_name} ma\'lumotlari yangilandi!')
            return redirect('employee_detail', id=employee.id)
        else:
            messages.error(request, 'Iltimos, xatolarni to\'g\'rilang.')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'employee_form_old.html', {
        'form': form,
        'employee': employee,
        'is_add': False,
    })